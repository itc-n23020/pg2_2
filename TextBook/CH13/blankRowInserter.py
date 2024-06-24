import sys
from openpyxl import load_workbook, Workbook

def insert_blank_rows(filename, N, M):
    # Excelファイルを読み込む
    wb = load_workbook(filename)
    ws = wb.active

    # 新しいワークブックとシートを作成
    new_wb = Workbook()
    new_ws = new_wb.active

    # 1からN行目までを新しいシートにコピー
    for row in ws.iter_rows(min_row=1, max_row=N, values_only=True):
        new_ws.append(row)

    # M行の空行を挿入
    for _ in range(M):
        new_ws.append([])

    # 残りの行を新しいシートにコピー
    for row in ws.iter_rows(min_row=N+1, values_only=True):
        new_ws.append(row)

    # 新しいファイル名を作成して保存
    new_filename = f'new_{filename}'
    new_wb.save(new_filename)
    print(f'{new_filename} が作成されました。')

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("使用法: python3 blankRowInserter.py <N> <M> <filename>")
        sys.exit(1)

    try:
        N = int(sys.argv[1])
        M = int(sys.argv[2])
        filename = sys.argv[3]
        insert_blank_rows(filename, N, M)
    except ValueError:
        print("NとMは整数である必要があります。")

