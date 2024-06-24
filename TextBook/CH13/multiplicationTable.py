import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

def create_multiplication_table(n):
    # 新しいExcelワークブックを作成
    wb = Workbook()
    ws = wb.active

    # 太字のフォントを定義
    bold_font = Font(bold=True)

    # ラベル行と列を作成
    for i in range(1, n+1):
        ws.cell(row=1, column=i+1, value=i).font = bold_font  # 第1行のラベル
        ws.cell(row=i+1, column=1, value=i).font = bold_font  # A列のラベル

    # 掛け算表を生成
    for i in range(1, n+1):
        for j in range(1, n+1):
            ws.cell(row=i+1, column=j+1, value=i*j)

    # ファイルに保存
    filename = f'multiplication_table_{n}x{n}.xlsx'
    wb.save(filename)
    print(f'{filename} が作成されました。')

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("使用法: python3 multiplicationTable.py <N>")
        sys.exit(1)

    try:
        n = int(sys.argv[1])
        create_multiplication_table(n)
    except ValueError:
        print("Nは整数である必要があります。")

