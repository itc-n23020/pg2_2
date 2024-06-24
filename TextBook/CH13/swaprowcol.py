import sys
from openpyxl import load_workbook, Workbook

def transpose_sheet(filename):
    # Excelファイルを読み込む
    wb = load_workbook(filename)
    ws = wb.active

    # データをリストのリストとして格納
    sheet_data = []
    for row in ws.iter_rows(values_only=True):
        sheet_data.append(list(row))

    # 新しいワークブックとシートを作成
    new_wb = Workbook()
    new_ws = new_wb.active

    # 行と列を入れ替えて新しいシートに書き込む
    for i in range(len(sheet_data)):
        for j in range(len(sheet_data[i])):
            new_ws.cell(row=j+1, column=i+1, value=sheet_data[i][j])

    # 新しいファイル名を作成して保存
    new_filename = f'transposed_{filename}'
    new_wb.save(new_filename)
    print(f'{new_filename} が作成されました。')

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("使用法: python3 transposeSheet.py <filename>")
        sys.exit(1)

    filename = sys.argv[1]
    transpose_sheet(filename)

